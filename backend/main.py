"""
세준 ERP - Phase 1~2 백엔드
기준정보(품목/거래처/공급사/창고) + 영업(주문/출고/반품/세금계산서) + 수금 + 출고 리포트

실행 방법은 README.md 참고.
"""

from datetime import date, datetime
from typing import Optional
from calendar import monthrange

from fastapi import FastAPI, HTTPException, Depends, UploadFile, File, Form
from openpyxl import load_workbook
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, model_validator
from sqlalchemy import create_engine, Column, Integer, String, Numeric, Boolean, Date, DateTime, ForeignKey, UniqueConstraint, text
from sqlalchemy.orm import declarative_base, sessionmaker, Session

# ---------------------------------------------------------------------------
# DB 설정 (SQLite 파일 하나로 동작. 별도 DB 서버 설치가 필요 없다.)
# ---------------------------------------------------------------------------
DATABASE_URL = "sqlite:///./sejun_erp.db"
engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()


# ---------------------------------------------------------------------------
# 테이블 정의
# ---------------------------------------------------------------------------
class Product(Base):
    __tablename__ = "products"

    id = Column(Integer, primary_key=True, index=True)

    sku = Column(String(50), unique=True, nullable=False)
    name = Column(String(200), nullable=False)
    spec = Column(String(100), nullable=True)  # 규격 추가

    unit_type = Column(String(20), nullable=False)

    units_per_multipack = Column(Integer, nullable=True)
    units_per_box = Column(Integer, nullable=True)

    box_price = Column(Numeric(12, 2), nullable=True)     # 박스 납품가
    multi_price = Column(Numeric(12, 2), nullable=True)   # 멀티 납품가
    unit_price = Column(Numeric(12, 2), nullable=True)    # 낱개 납품가

    cost_price = Column(Numeric(12, 2), nullable=True)

    supply_price = Column(Numeric(12, 2), nullable=True)

    stock_quantity = Column(Integer, default=0)  # 현재고(실사 기준 수동 입력, 자동 증감 아님)

    is_active = Column(Boolean, default=True)
    created_at = Column(DateTime, default=datetime.utcnow)


class PriceGroup(Base):
    """거래처그룹별 단가표. 여기 없는(=지정 안 된) 거래처는 품목 자체의 기본가(여수할인마트가)를 쓴다."""
    __tablename__ = "price_groups"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String(100), unique=True, nullable=False)


class ProductPrice(Base):
    """품목 x 가격그룹 조합별 박스/멀티/낱개 단가. 기본(여수할인마트) 그룹은 여기 없이 Product 자체 필드를 쓴다."""
    __tablename__ = "product_prices"
    __table_args__ = (UniqueConstraint("product_id", "price_group_id", name="uq_product_price_group"),)
    id = Column(Integer, primary_key=True, index=True)
    product_id = Column(Integer, ForeignKey("products.id"), nullable=False)
    price_group_id = Column(Integer, ForeignKey("price_groups.id"), nullable=False)
    box_price = Column(Numeric(12, 2), default=0)
    multi_price = Column(Numeric(12, 2), default=0)
    unit_price = Column(Numeric(12, 2), default=0)


class Customer(Base):
    __tablename__ = "customers"
    id = Column(Integer, primary_key=True, index=True)
    customer_code = Column(String(50), nullable=True)
    name = Column(String(200), nullable=False)
    biz_reg_no = Column(String(20), nullable=True)
    contact_name = Column(String(50), nullable=True)
    phone = Column(String(20), nullable=True)
    address = Column(String(255), nullable=True)
    price_group_id = Column(Integer, ForeignKey("price_groups.id"), nullable=True)  # 비어있으면 기본(여수할인마트)가
    outstanding_balance = Column(Numeric(12, 2), default=0)  # 현재 미수금 잔액
    created_at = Column(DateTime, default=datetime.utcnow)


class Supplier(Base):
    __tablename__ = "suppliers"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String(200), nullable=False)
    biz_reg_no = Column(String(20), nullable=True)
    contact_name = Column(String(50), nullable=True)
    phone = Column(String(20), nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)


class Warehouse(Base):
    __tablename__ = "warehouses"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String(100), nullable=False)
    address = Column(String(255), nullable=True)


class Purchase(Base):
    """매입 기록. 특정 공급사에 연결하지 않고 그날의 매입금액만 기록한다."""
    __tablename__ = "purchases"
    id = Column(Integer, primary_key=True, index=True)
    amount = Column(Numeric(12, 2), nullable=False)
    purchased_at = Column(Date, nullable=False)
    memo = Column(String(255), nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)


class InventoryValue(Base):
    """월별 재고금액을 직접 입력해서 관리한다(품목별 자동 계산이 아님)."""
    __tablename__ = "inventory_values"
    id = Column(Integer, primary_key=True, index=True)
    month = Column(String(7), nullable=False)  # 'YYYY-MM' 형식
    amount = Column(Numeric(14, 2), nullable=False)
    memo = Column(String(255), nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)


class Collection(Base):
    __tablename__ = "collections"
    id = Column(Integer, primary_key=True, index=True)
    customer_id = Column(Integer, ForeignKey("customers.id"), nullable=False)
    amount = Column(Numeric(12, 2), nullable=False)
    collected_at = Column(Date, nullable=False)
    method = Column(String(20), nullable=True)  # 현금 | 계좌이체 | 카드 | 어음 등
    memo = Column(String(255), nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)


class SalesOrder(Base):
    __tablename__ = "sales_orders"
    id = Column(Integer, primary_key=True, index=True)
    customer_id = Column(Integer, ForeignKey("customers.id"), nullable=False)
    order_date = Column(Date, nullable=False)
    status = Column(String(20), default="주문접수")  # 주문접수 | 출고완료
    supply_amount = Column(Numeric(12, 2), default=0)  # 공급가액 합계(VAT 제외)
    vat_amount = Column(Numeric(12, 2), default=0)     # 부가세(공급가액의 10%, 단순 계산)
    total_amount = Column(Numeric(12, 2), default=0)   # 합계금액(공급가액+부가세)
    cost_amount = Column(Numeric(12, 2), default=0)     # 원가 합계(마진 계산용)
    shipped_at = Column(DateTime, nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)


class SalesOrderItem(Base):
    __tablename__ = "sales_order_items"
    id = Column(Integer, primary_key=True, index=True)
    order_id = Column(Integer, ForeignKey("sales_orders.id"), nullable=False)
    product_id = Column(Integer, ForeignKey("products.id"), nullable=False)
    product_name = Column(String(200), nullable=False)  # 주문 시점 품명 스냅샷
    unit_type = Column(String(20), nullable=True)  # 박스 | 멀티팩 | 낱개 (기존 데이터는 비어있을 수 있음)
    quantity = Column(Integer, nullable=False)
    unit_price = Column(Numeric(12, 2), nullable=False)  # 주문 시점 판가 스냅샷(VAT 제외)
    unit_cost = Column(Numeric(12, 2), nullable=False, default=0)  # 주문 시점 원가 스냅샷
    line_amount = Column(Numeric(12, 2), nullable=False)  # quantity * unit_price


class Return(Base):
    __tablename__ = "returns"
    id = Column(Integer, primary_key=True, index=True)
    order_item_id = Column(Integer, ForeignKey("sales_order_items.id"), nullable=False)
    quantity = Column(Integer, nullable=False)
    reason = Column(String(255), nullable=True)
    returned_at = Column(Date, nullable=False)
    refund_amount = Column(Numeric(12, 2), nullable=False)  # VAT 포함 환불액
    created_at = Column(DateTime, default=datetime.utcnow)


class TaxInvoice(Base):
    __tablename__ = "tax_invoices"
    id = Column(Integer, primary_key=True, index=True)
    order_id = Column(Integer, ForeignKey("sales_orders.id"), nullable=False, unique=True)
    invoice_no = Column(String(50), nullable=False)
    issue_date = Column(Date, nullable=False)
    supply_amount = Column(Numeric(12, 2), nullable=False)
    vat_amount = Column(Numeric(12, 2), nullable=False)
    total_amount = Column(Numeric(12, 2), nullable=False)
    created_at = Column(DateTime, default=datetime.utcnow)


Base.metadata.create_all(bind=engine)


def _ensure_column(table: str, column: str, coltype: str):
    """이미 존재하는 로컬 DB 파일에 새 컬럼을 안전하게 추가한다(기존 데이터는 보존)."""
    with engine.connect() as conn:
        cols = [row[1] for row in conn.execute(text(f"PRAGMA table_info({table})"))]
        if column not in cols:
            conn.execute(text(f"ALTER TABLE {table} ADD COLUMN {column} {coltype}"))
            conn.commit()


_ensure_column("customers", "price_group_id", "INTEGER")
_ensure_column("sales_order_items", "unit_type", "VARCHAR(20)")
_ensure_column("products", "stock_quantity", "INTEGER")


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# ---------------------------------------------------------------------------
# Pydantic 스키마
# ---------------------------------------------------------------------------

class ProductIn(BaseModel):
    sku: str
    name: str
    spec: Optional[str] = None

    unit_type: str

    units_per_multipack: Optional[int] = None
    units_per_box: Optional[int] = None

    box_price: Optional[float] = None
    multi_price: Optional[float] = None
    unit_price: Optional[float] = None

    cost_price: Optional[float] = None
    supply_price: Optional[float] = None

    stock_quantity: int = 0

    is_active: bool = True

class ProductOut(BaseModel):
    id: int

    sku: str
    name: str
    spec: Optional[str] = None

    unit_type: str

    units_per_multipack: Optional[int] = None
    units_per_box: Optional[int] = None

    box_price: Optional[float] = None
    multi_price: Optional[float] = None
    unit_price: Optional[float] = None

    cost_price: Optional[float] = None
    supply_price: Optional[float] = None

    stock_quantity: int = 0

    margin_rate: Optional[float] = None

    is_active: bool

    class Config:
        from_attributes = True

    @model_validator(mode="after")
    def _compute_margin(self):
        if self.supply_price and self.cost_price is not None:
            self.margin_rate = round(
                (self.supply_price - self.cost_price)
                / self.supply_price * 100,
                2
            )
        return self

class CustomerIn(BaseModel):
    customer_code: Optional[str] = None
    name: str
    biz_reg_no: Optional[str] = None
    contact_name: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    price_group_id: Optional[int] = None
    outstanding_balance: float = 0


class CustomerOut(CustomerIn):
    id: int
    price_group_name: Optional[str] = None


class PriceGroupOut(BaseModel):
    id: int
    name: str


class ProductPriceOut(BaseModel):
    product_id: int
    product_name: Optional[str] = None
    price_group_id: int
    price_group_name: Optional[str] = None
    box_price: float = 0
    multi_price: float = 0
    unit_price: float = 0


class SupplierIn(BaseModel):
    name: str
    biz_reg_no: Optional[str] = None
    contact_name: Optional[str] = None
    phone: Optional[str] = None


class SupplierOut(SupplierIn):
    id: int

    class Config:
        from_attributes = True


class WarehouseIn(BaseModel):
    name: str
    address: Optional[str] = None


class WarehouseOut(WarehouseIn):
    id: int

    class Config:
        from_attributes = True


class CollectionIn(BaseModel):
    customer_id: int
    amount: float
    collected_at: date
    method: Optional[str] = None
    memo: Optional[str] = None


class CollectionOut(CollectionIn):
    id: int
    customer_name: Optional[str] = None

    class Config:
        from_attributes = True


class PurchaseIn(BaseModel):
    amount: float
    purchased_at: date
    memo: Optional[str] = None


class PurchaseOut(PurchaseIn):
    id: int

    class Config:
        from_attributes = True


class InventoryValueIn(BaseModel):
    month: str  # 'YYYY-MM'
    amount: float
    memo: Optional[str] = None


class InventoryValueOut(InventoryValueIn):
    id: int

    class Config:
        from_attributes = True


class OrderItemIn(BaseModel):
    product_id: int
    unit_type: str  # 박스 | 멀티팩 | 낱개
    quantity: int


class OrderItemOut(BaseModel):
    id: int
    product_id: int
    product_name: str
    unit_type: Optional[str] = None
    quantity: int
    returned_quantity: int = 0
    unit_price: float
    unit_cost: float
    line_amount: float

    class Config:
        from_attributes = True


class CustomerSalesHistoryItemOut(BaseModel):
    order_id: int
    order_date: date
    status: str
    product_name: str
    unit_type: Optional[str] = None
    quantity: int
    unit_price: float
    line_amount: float


class SalesOrderIn(BaseModel):
    customer_id: int
    order_date: date
    items: list[OrderItemIn]


class SalesOrderOut(BaseModel):
    id: int
    customer_id: int
    customer_name: Optional[str] = None
    order_date: date
    status: str
    supply_amount: float
    vat_amount: float
    total_amount: float
    cost_amount: float
    margin_amount: float = 0
    margin_rate: Optional[float] = None
    shipped_at: Optional[datetime] = None
    items: list[OrderItemOut] = []
    label: Optional[str] = None  # 프론트 select 표시용 요약 문자열

    class Config:
        from_attributes = True


class ReturnIn(BaseModel):
    order_item_id: int
    quantity: int
    reason: Optional[str] = None
    returned_at: date


class ReturnOut(BaseModel):
    id: int
    order_item_id: int
    product_name: Optional[str] = None
    customer_name: Optional[str] = None
    quantity: int
    reason: Optional[str] = None
    returned_at: date
    refund_amount: float

    class Config:
        from_attributes = True


class TaxInvoiceIn(BaseModel):
    order_id: int
    invoice_no: str
    issue_date: date


class TaxInvoiceOut(BaseModel):
    id: int
    order_id: int
    customer_name: Optional[str] = None
    invoice_no: str
    issue_date: date
    supply_amount: float
    vat_amount: float
    total_amount: float

    class Config:
        from_attributes = True


class ShipmentReportOut(BaseModel):
    period: str
    order_count: int
    supply_amount: float
    cost_amount: float
    margin_amount: float
    margin_rate: Optional[float] = None


# ---------------------------------------------------------------------------
# FastAPI 앱
# ---------------------------------------------------------------------------
app = FastAPI(title="Sejun ERP API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # 로컬/사무실 내부망 사용 목적이라 전체 허용.
    allow_methods=["*"],
    allow_headers=["*"],
)

def _crud_router(model, in_schema, out_schema, prefix: str, name: str):
    """품목/거래처/공급사/창고는 구조가 비슷하므로 CRUD 라우트를 공통 함수로 생성한다."""

    @app.get(f"/{prefix}", response_model=list[out_schema], tags=[name])
    def list_items(db: Session = Depends(get_db)):
        return db.query(model).all()

    @app.get(f"/{prefix}/{{item_id}}", response_model=out_schema, tags=[name])
    def get_item(item_id: int, db: Session = Depends(get_db)):
        item = db.query(model).filter(model.id == item_id).first()
        if not item:
            raise HTTPException(status_code=404, detail=f"{name} not found")
        return item

    @app.post(f"/{prefix}", response_model=out_schema, tags=[name])
    def create_item(payload: in_schema, db: Session = Depends(get_db)):
        item = model(**payload.model_dump())
        db.add(item)
        db.commit()
        db.refresh(item)
        return item

    @app.put(f"/{prefix}/{{item_id}}", response_model=out_schema, tags=[name])
    def update_item(item_id: int, payload: in_schema, db: Session = Depends(get_db)):
        item = db.query(model).filter(model.id == item_id).first()
        if not item:
            raise HTTPException(status_code=404, detail=f"{name} not found")
        for key, value in payload.model_dump().items():
            setattr(item, key, value)
        db.commit()
        db.refresh(item)
        return item

    @app.delete(f"/{prefix}/{{item_id}}", tags=[name])
    def delete_item(item_id: int, db: Session = Depends(get_db)):
        item = db.query(model).filter(model.id == item_id).first()
        if not item:
            raise HTTPException(status_code=404, detail=f"{name} not found")
        db.delete(item)
        db.commit()
        return {"ok": True}


_crud_router(Product, ProductIn, ProductOut, "products", "품목")
_crud_router(Supplier, SupplierIn, SupplierOut, "suppliers", "공급사")
_crud_router(Warehouse, WarehouseIn, WarehouseOut, "warehouses", "창고")


# ---------------------------------------------------------------------------
# 가격그룹 라우트
# ---------------------------------------------------------------------------
@app.get("/price-groups", response_model=list[PriceGroupOut], tags=["가격그룹"])
def list_price_groups(db: Session = Depends(get_db)):
    return [PriceGroupOut(id=g.id, name=g.name) for g in db.query(PriceGroup).all()]


@app.get("/product-prices", response_model=list[ProductPriceOut], tags=["가격그룹"])
def list_product_prices_by_group(price_group_id: int, db: Session = Depends(get_db)):
    rows = db.query(ProductPrice).filter(ProductPrice.price_group_id == price_group_id).all()
    group = db.query(PriceGroup).filter(PriceGroup.id == price_group_id).first()
    result = []
    for r in rows:
        product = db.query(Product).filter(Product.id == r.product_id).first()
        result.append(ProductPriceOut(
            product_id=r.product_id, product_name=product.name if product else None,
            price_group_id=price_group_id, price_group_name=group.name if group else None,
            box_price=float(r.box_price or 0), multi_price=float(r.multi_price or 0),
            unit_price=float(r.unit_price or 0),
        ))
    return result


# ---------------------------------------------------------------------------
# 거래처 라우트 (가격그룹명 조인을 위해 공통 CRUD 대신 개별 작성)
# ---------------------------------------------------------------------------
def _customer_to_out(c: Customer, db: Session) -> CustomerOut:
    group_name = None
    if c.price_group_id:
        g = db.query(PriceGroup).filter(PriceGroup.id == c.price_group_id).first()
        group_name = g.name if g else None
    else:
        group_name = "기본(여수할인마트)"
    return CustomerOut(
        id=c.id, customer_code=c.customer_code, name=c.name, biz_reg_no=c.biz_reg_no,
        contact_name=c.contact_name, phone=c.phone, address=c.address,
        price_group_id=c.price_group_id, price_group_name=group_name,
        outstanding_balance=float(c.outstanding_balance or 0),
    )


@app.get("/customers", response_model=list[CustomerOut], tags=["거래처"])
def list_customers(db: Session = Depends(get_db)):
    return [_customer_to_out(c, db) for c in db.query(Customer).order_by(Customer.id).all()]


@app.post("/customers", response_model=CustomerOut, tags=["거래처"])
def create_customer(payload: CustomerIn, db: Session = Depends(get_db)):
    c = Customer(**payload.model_dump())
    db.add(c)
    db.commit()
    db.refresh(c)
    return _customer_to_out(c, db)


@app.put("/customers/{item_id}", response_model=CustomerOut, tags=["거래처"])
def update_customer(item_id: int, payload: CustomerIn, db: Session = Depends(get_db)):
    c = db.query(Customer).filter(Customer.id == item_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="거래처를 찾을 수 없습니다")
    for key, value in payload.model_dump().items():
        setattr(c, key, value)
    db.commit()
    db.refresh(c)
    return _customer_to_out(c, db)


@app.delete("/customers/{item_id}", tags=["거래처"])
def delete_customer(item_id: int, db: Session = Depends(get_db)):
    c = db.query(Customer).filter(Customer.id == item_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="거래처를 찾을 수 없습니다")
    db.delete(c)
    db.commit()
    return {"ok": True}


# ---------------------------------------------------------------------------
# 수금 라우트
# ---------------------------------------------------------------------------
def _collection_to_out(c: Collection, db: Session) -> CollectionOut:
    customer = db.query(Customer).filter(Customer.id == c.customer_id).first()
    return CollectionOut(
        id=c.id, customer_id=c.customer_id,
        customer_name=customer.name if customer else None,
        amount=float(c.amount), collected_at=c.collected_at,
        method=c.method, memo=c.memo,
    )


@app.get("/collections", response_model=list[CollectionOut], tags=["수금"])
def list_collections(db: Session = Depends(get_db)):
    items = db.query(Collection).order_by(Collection.collected_at.desc()).all()
    return [_collection_to_out(c, db) for c in items]


@app.post("/collections", response_model=CollectionOut, tags=["수금"])
def create_collection(payload: CollectionIn, db: Session = Depends(get_db)):
    customer = db.query(Customer).filter(Customer.id == payload.customer_id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="거래처를 찾을 수 없습니다")

    item = Collection(**payload.model_dump())
    db.add(item)
    customer.outstanding_balance = float(customer.outstanding_balance or 0) - payload.amount

    db.commit()
    db.refresh(item)
    return _collection_to_out(item, db)


@app.delete("/collections/{item_id}", tags=["수금"])
def delete_collection(item_id: int, db: Session = Depends(get_db)):
    item = db.query(Collection).filter(Collection.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="수금 내역을 찾을 수 없습니다")

    customer = db.query(Customer).filter(Customer.id == item.customer_id).first()
    if customer:
        customer.outstanding_balance = float(customer.outstanding_balance or 0) + float(item.amount)

    db.delete(item)
    db.commit()
    return {"ok": True}


# ---------------------------------------------------------------------------
# 매입 / 재고금액 라우트 (둘 다 날짜(또는 월) + 금액만 있는 단순 기록이라 공통 CRUD 사용)
# ---------------------------------------------------------------------------
_crud_router(Purchase, PurchaseIn, PurchaseOut, "purchases", "매입")
_crud_router(InventoryValue, InventoryValueIn, InventoryValueOut, "inventory-values", "재고금액")


# ---------------------------------------------------------------------------
# 엑셀 업로드 (거래처 정보 갱신) - 업로드해주신 버전 그대로 유지
# ---------------------------------------------------------------------------
@app.post("/admin/import-initial-catalog", tags=["엑셀"])
def import_initial_catalog(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):

    wb = load_workbook(file.file)
    ws = wb.active

    count = 0

    print("엑셀 업로드 실행됨")

    for row in ws.iter_rows(min_row=3, values_only=True):

        if not row[0]:
            continue

        sku = str(row[0]).strip()
        name = row[1]
        spec = row[2]

        box_price = row[3] or 0
        multi_price = row[4] or 0
        unit_price = row[5] or 0

        print("가격확인:", sku, box_price, multi_price, unit_price)


        product = db.query(Product).filter(
            Product.sku == sku
        ).first()


        if product:

            product.name = name
            product.spec = spec

            product.box_price = box_price
            product.multi_price = multi_price
            product.unit_price = unit_price

            # 화면 판가 표시용
            product.supply_price = box_price


        else:

            product = Product(
                sku=sku,
                name=name,
                spec=spec,

                unit_type="박스",

                box_price=box_price,
                multi_price=multi_price,
                unit_price=unit_price,

                supply_price=box_price
            )

            db.add(product)


        count += 1


    db.commit()

    print("저장확인:", product.sku, product.box_price, product.supply_price)

    print("현재 상품 개수:", db.query(Product).count())


    return {
    "업데이트건수": count
}


@app.post("/admin/import-group-price", tags=["엑셀"])
def import_group_price(
    file: UploadFile = File(...),
    group_name: str = Form(...),
    apply_to_customers: str = Form(...),  # 쉼표로 구분된 거래처명 목록
    db: Session = Depends(get_db),
):
    """기본(여수할인마트)가와 다른 가격그룹(예: 여천식자재, 푸드원)을 업로드하고,
    지정한 거래처들의 가격그룹을 이 그룹으로 일괄 배정한다."""

    group = db.query(PriceGroup).filter(PriceGroup.name == group_name).first()
    if not group:
        group = PriceGroup(name=group_name)
        db.add(group)
        db.flush()

    wb = load_workbook(file.file)
    ws = wb.active

    price_count = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[0]:
            continue
        sku = str(row[0]).strip()
        name = row[1]
        spec = row[2]
        box_price = row[3] or 0
        multi_price = row[4] or 0
        unit_price = row[5] or 0

        product = db.query(Product).filter(Product.sku == sku).first()
        if not product:
            # 이 그룹에만 있는 품목이면 기본 정보로 새로 등록한다(기본가는 비워둠).
            product = Product(sku=sku, name=name, spec=spec, unit_type="박스")
            db.add(product)
            db.flush()

        price_row = db.query(ProductPrice).filter(
            ProductPrice.product_id == product.id, ProductPrice.price_group_id == group.id
        ).first()
        if not price_row:
            price_row = ProductPrice(product_id=product.id, price_group_id=group.id)
            db.add(price_row)
        price_row.box_price = box_price
        price_row.multi_price = multi_price
        price_row.unit_price = unit_price
        price_count += 1

    customer_count = 0
    for raw_name in apply_to_customers.split(","):
        cust_name = raw_name.strip()
        if not cust_name:
            continue
        customer = db.query(Customer).filter(Customer.name == cust_name).first()
        if not customer:
            customer = Customer(name=cust_name, price_group_id=group.id)
            db.add(customer)
        else:
            customer.price_group_id = group.id
        customer_count += 1

    db.commit()

    return {
        "가격그룹": group_name,
        "가격등록건수": price_count,
        "배정된거래처수": customer_count,
    }


@app.post("/admin/assign-price-group", tags=["가격그룹"])
def assign_price_group(
    group_name: str = Form(...),
    customer_names: str = Form(...),  # 쉼표로 구분된 거래처명 목록
    db: Session = Depends(get_db),
):
    """가격표 파일 없이, 이미 존재하는 가격그룹에 거래처만 배정하거나 재배정한다.
    거래처 이름이 정확히 일치해야 한다(거래처 탭에서 정확한 이름을 확인하고 사용)."""

    group = db.query(PriceGroup).filter(PriceGroup.name == group_name).first()
    if not group:
        group = PriceGroup(name=group_name)
        db.add(group)
        db.flush()

    assigned, not_found = [], []
    for raw_name in customer_names.split(","):
        cust_name = raw_name.strip()
        if not cust_name:
            continue
        customer = db.query(Customer).filter(Customer.name == cust_name).first()
        if not customer:
            not_found.append(cust_name)
            continue
        customer.price_group_id = group.id
        assigned.append(cust_name)

    db.commit()

    return {
        "가격그룹": group_name,
        "배정됨": assigned,
        "찾지못함": not_found,
    }


@app.post("/admin/import-customer-list", tags=["엑셀"])
def import_customer_list(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):

    wb = load_workbook(file.file)
    ws = wb.active

    updated = 0
    created = 0

    for row in ws.iter_rows(min_row=2, values_only=True):

        if not row[1]:
            continue

        customer_code = str(row[1]).strip()
        name = row[2]
        biz_reg_no = row[6]
        contact_name = row[9]
        phone = row[10]
        address = row[11]

        customer = db.query(Customer).filter(
            Customer.customer_code == customer_code
        ).first()

        if not customer and name:
            # 거래처코드로 못 찾았다면, 이름이 같은 거래처가 있는지 한 번 더 확인한다.
            # (가격표 업로드로 자동 생성된 거래처는 코드가 비어있는 채로 만들어지기 때문)
            candidate = db.query(Customer).filter(Customer.name == str(name).strip()).first()
            if candidate:
                customer = candidate
                customer.customer_code = customer_code

        if customer:
            customer.name = name or customer.name
            customer.biz_reg_no = biz_reg_no
            customer.contact_name = contact_name
            customer.phone = phone
            customer.address = address
            updated += 1
        else:
            # 기존에는 매칭되는 거래처가 없으면 건너뛰기만 했는데,
            # 그러면 신규 거래처 목록 자체를 올릴 때 아무것도 등록되지 않는 문제가 있었다.
            # 이제는 없으면 새로 만든다.
            customer = Customer(
                customer_code=customer_code,
                name=name or customer_code,
                biz_reg_no=biz_reg_no,
                contact_name=contact_name,
                phone=phone,
                address=address,
            )
            db.add(customer)
            created += 1

    db.commit()

    return {
        "신규등록건수": created,
        "업데이트건수": updated,
    }


# 가격표 업로드 때 자동 생성된 짧은 이름 거래처와, 정식 거래처목록의 정식 명칭이 서로 달라
# 중복으로 쌓이는 문제가 있었다. 실제로 확인된 짝을 여기에 등록해두고 병합에 사용한다.
CUSTOMER_NAME_ALIASES = {
    "주식회사푸드원": ["푸드원"],
    "(유)여천식자재": ["여천식자재"],
    "여수웅천식자재 주식회사": ["웅천식자재"],
    "웅천식자재도매센터(주)": ["웅천식자재도매센터"],
}


@app.post("/admin/import-customer-full", tags=["엑셀"])
def import_customer_full(
    customer_list_file: UploadFile = File(...),
    receivables_file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """거래처목록(사업자번호/담당자/전화번호/주소)과 거래처미수현황(당월미수)을
    함께 업로드해 거래처 하나당 하나의 레코드로 병합하고 미수금까지 반영한다."""

    # 1) 거래처목록 반영
    wb1 = load_workbook(customer_list_file.file)
    ws1 = wb1.active
    created, updated = 0, 0

    for row in ws1.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:
            continue
        code = str(row[1]).strip()
        name = str(row[2]).strip() if row[2] else None
        biz_reg_no = str(row[6]).strip() if row[6] else None
        rep_name = row[8] or None
        contact_name = (row[9] or rep_name) or None
        phone = (row[10] or row[7]) or None
        address = str(row[11]).strip() if row[11] else None

        customer = db.query(Customer).filter(Customer.customer_code == code).first()
        if not customer and name:
            customer = db.query(Customer).filter(Customer.name == name).first()
        if not customer and name in CUSTOMER_NAME_ALIASES:
            for alias in CUSTOMER_NAME_ALIASES[name]:
                customer = db.query(Customer).filter(Customer.name == alias).first()
                if customer:
                    break

        if customer:
            customer.customer_code = code
            customer.name = name or customer.name
            customer.biz_reg_no = biz_reg_no
            customer.contact_name = contact_name
            customer.phone = phone
            customer.address = address
            updated += 1
        else:
            customer = Customer(
                customer_code=code, name=name or code, biz_reg_no=biz_reg_no,
                contact_name=contact_name, phone=phone, address=address,
            )
            db.add(customer)
            created += 1

    db.flush()

    # 2) 짧은 이름으로 남아있는 중복 거래처를 정식 명칭 거래처로 병합
    merged = 0
    for official_name, aliases in CUSTOMER_NAME_ALIASES.items():
        official = db.query(Customer).filter(Customer.name == official_name).first()
        if not official:
            continue
        for alias in aliases:
            dup = db.query(Customer).filter(Customer.name == alias).first()
            if not dup or dup.id == official.id:
                continue
            if official.price_group_id is None and dup.price_group_id is not None:
                official.price_group_id = dup.price_group_id
            has_refs = (
                db.query(SalesOrder).filter(SalesOrder.customer_id == dup.id).first()
                or db.query(Collection).filter(Collection.customer_id == dup.id).first()
            )
            if has_refs:
                # 이미 주문/수금 기록이 걸려있으면 삭제 대신 이름만 구분해두고 남긴다.
                dup.name = f"{alias}(중복-사용안함)"
            else:
                db.delete(dup)
            merged += 1

    # 3) 거래처미수현황 반영 (당월미수 -> 미수금)
    wb2 = load_workbook(receivables_file.file)
    ws2 = wb2.active
    balance_updated = 0

    for row in ws2.iter_rows(min_row=3, values_only=True):
        if not row or row[0] is None:
            continue
        raw_code = row[0]
        code = str(int(raw_code)).strip() if isinstance(raw_code, float) else str(raw_code).strip()
        outstanding = row[6] or 0

        customer = db.query(Customer).filter(Customer.customer_code == code).first()
        if not customer and row[1]:
            customer = db.query(Customer).filter(Customer.name == str(row[1]).strip()).first()
        if customer:
            customer.outstanding_balance = float(outstanding)
            balance_updated += 1

    db.commit()

    return {
        "거래처_신규등록": created,
        "거래처_정보갱신": updated,
        "중복병합": merged,
        "미수금_반영건수": balance_updated,
    }


# ---------------------------------------------------------------------------
# 주문 라우트
# ---------------------------------------------------------------------------
def _order_to_out(order: SalesOrder, db: Session) -> SalesOrderOut:
    customer = db.query(Customer).filter(Customer.id == order.customer_id).first()
    raw_items = db.query(SalesOrderItem).filter(SalesOrderItem.order_id == order.id).all()
    item_outs = []
    for it in raw_items:
        returned_sum = sum(
            r.quantity for r in db.query(Return).filter(Return.order_item_id == it.id).all()
        )
        item_outs.append(OrderItemOut(
            id=it.id, product_id=it.product_id, product_name=it.product_name,
            unit_type=it.unit_type, quantity=it.quantity, returned_quantity=returned_sum,
            unit_price=float(it.unit_price), unit_cost=float(it.unit_cost),
            line_amount=float(it.line_amount),
        ))
    supply = float(order.supply_amount)
    cost = float(order.cost_amount)
    margin_amount = supply - cost
    margin_rate = round(margin_amount / supply * 100, 2) if supply else None
    label = f"#{order.id} {order.order_date} / {customer.name if customer else '-'} / {int(order.total_amount):,}원"
    return SalesOrderOut(
        id=order.id, customer_id=order.customer_id,
        customer_name=customer.name if customer else None,
        order_date=order.order_date, status=order.status,
        supply_amount=supply, vat_amount=float(order.vat_amount),
        total_amount=float(order.total_amount), cost_amount=cost,
        margin_amount=margin_amount, margin_rate=margin_rate,
        shipped_at=order.shipped_at, items=item_outs, label=label,
    )


@app.get("/sales-orders", response_model=list[SalesOrderOut], tags=["주문"])
def list_orders(db: Session = Depends(get_db)):
    orders = db.query(SalesOrder).order_by(SalesOrder.id.desc()).all()
    return [_order_to_out(o, db) for o in orders]


@app.get("/customers/{customer_id}/sales-history", response_model=list[CustomerSalesHistoryItemOut], tags=["거래처"])
def customer_sales_history(customer_id: int, db: Session = Depends(get_db)):
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="거래처를 찾을 수 없습니다")

    rows = (
        db.query(SalesOrderItem, SalesOrder)
        .join(SalesOrder, SalesOrder.id == SalesOrderItem.order_id)
        .filter(SalesOrder.customer_id == customer_id)
        .order_by(SalesOrder.order_date.desc(), SalesOrder.id.desc())
        .all()
    )
    return [
        CustomerSalesHistoryItemOut(
            order_id=order.id, order_date=order.order_date, status=order.status,
            product_name=item.product_name, unit_type=item.unit_type,
            quantity=item.quantity, unit_price=float(item.unit_price),
            line_amount=float(item.line_amount),
        )
        for item, order in rows
    ]


UNIT_PRICE_FIELD = {"박스": "box_price", "멀티팩": "multi_price", "낱개": "unit_price"}


@app.post("/sales-orders", response_model=SalesOrderOut, tags=["주문"])
def create_order(payload: SalesOrderIn, db: Session = Depends(get_db)):
    customer = db.query(Customer).filter(Customer.id == payload.customer_id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="거래처를 찾을 수 없습니다")
    if not payload.items:
        raise HTTPException(status_code=400, detail="주문 항목이 최소 1개 필요합니다")

    order = SalesOrder(customer_id=payload.customer_id, order_date=payload.order_date, status="주문접수")
    db.add(order)
    db.flush()

    supply_total = 0.0
    cost_total = 0.0
    for line in payload.items:
        if line.quantity <= 0:
            raise HTTPException(status_code=400, detail="수량은 1 이상이어야 합니다")
        if line.unit_type not in UNIT_PRICE_FIELD:
            raise HTTPException(status_code=400, detail="단위는 박스/멀티팩/낱개 중 하나여야 합니다")
        product = db.query(Product).filter(Product.id == line.product_id).first()
        if not product:
            raise HTTPException(status_code=404, detail=f"품목 ID {line.product_id}를 찾을 수 없습니다")

        price_field = UNIT_PRICE_FIELD[line.unit_type]
        if customer.price_group_id:
            price_row = db.query(ProductPrice).filter(
                ProductPrice.product_id == product.id, ProductPrice.price_group_id == customer.price_group_id
            ).first()
            if not price_row:
                raise HTTPException(status_code=404, detail=f"{product.name}의 이 거래처 가격그룹 단가가 없습니다")
            unit_price = float(getattr(price_row, price_field) or 0)
        else:
            unit_price = float(getattr(product, price_field) or 0)

        if unit_price == 0:
            raise HTTPException(status_code=400, detail=f"{product.name}은(는) {line.unit_type} 단위로 판매하지 않습니다")

        unit_cost = float(product.cost_price or 0)
        line_amount = unit_price * line.quantity
        supply_total += line_amount
        cost_total += unit_cost * line.quantity
        db.add(SalesOrderItem(
            order_id=order.id, product_id=product.id, product_name=product.name,
            unit_type=line.unit_type, quantity=line.quantity,
            unit_price=unit_price, unit_cost=unit_cost, line_amount=line_amount,
        ))

    vat_total = round(supply_total * 0.1)  # 단순화된 VAT 계산(10%, 주문 합계 기준 반올림)
    order.supply_amount = supply_total
    order.vat_amount = vat_total
    order.total_amount = supply_total + vat_total
    order.cost_amount = cost_total

    # 주문 등록 시점에 미수금을 반영한다.
    customer.outstanding_balance = float(customer.outstanding_balance or 0) + order.total_amount

    db.commit()
    db.refresh(order)
    return _order_to_out(order, db)


@app.patch("/sales-orders/{order_id}/ship", response_model=SalesOrderOut, tags=["주문"])
def ship_order(order_id: int, db: Session = Depends(get_db)):
    order = db.query(SalesOrder).filter(SalesOrder.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="주문을 찾을 수 없습니다")
    if order.status == "출고완료":
        raise HTTPException(status_code=400, detail="이미 출고 처리된 주문입니다")
    order.status = "출고완료"
    order.shipped_at = datetime.utcnow()
    db.commit()
    db.refresh(order)
    return _order_to_out(order, db)


# ---------------------------------------------------------------------------
# 반품 라우트
# ---------------------------------------------------------------------------
def _return_to_out(r: Return, db: Session) -> ReturnOut:
    item = db.query(SalesOrderItem).filter(SalesOrderItem.id == r.order_item_id).first()
    product_name = item.product_name if item else None
    customer_name = None
    if item:
        order = db.query(SalesOrder).filter(SalesOrder.id == item.order_id).first()
        if order:
            customer = db.query(Customer).filter(Customer.id == order.customer_id).first()
            customer_name = customer.name if customer else None
    return ReturnOut(
        id=r.id, order_item_id=r.order_item_id, product_name=product_name,
        customer_name=customer_name, quantity=r.quantity, reason=r.reason,
        returned_at=r.returned_at, refund_amount=float(r.refund_amount),
    )


@app.get("/returns", response_model=list[ReturnOut], tags=["반품"])
def list_returns(db: Session = Depends(get_db)):
    items = db.query(Return).order_by(Return.id.desc()).all()
    return [_return_to_out(r, db) for r in items]


@app.post("/returns", response_model=ReturnOut, tags=["반품"])
def create_return(payload: ReturnIn, db: Session = Depends(get_db)):
    item = db.query(SalesOrderItem).filter(SalesOrderItem.id == payload.order_item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="주문 항목을 찾을 수 없습니다")
    if payload.quantity <= 0:
        raise HTTPException(status_code=400, detail="반품 수량은 1 이상이어야 합니다")

    returned_sum = sum(
        r.quantity for r in db.query(Return).filter(Return.order_item_id == item.id).all()
    )
    remaining = item.quantity - returned_sum
    if payload.quantity > remaining:
        raise HTTPException(status_code=400, detail=f"반품 가능 수량({remaining}개)을 초과했습니다")

    order = db.query(SalesOrder).filter(SalesOrder.id == item.order_id).first()
    customer = db.query(Customer).filter(Customer.id == order.customer_id).first() if order else None

    # 환불액은 VAT 포함 금액으로 계산한다(단가는 VAT 제외 스냅샷이므로 1.1배 적용).
    refund_amount = float(item.unit_price) * payload.quantity * 1.1

    rec = Return(
        order_item_id=item.id, quantity=payload.quantity, reason=payload.reason,
        returned_at=payload.returned_at, refund_amount=refund_amount,
    )
    db.add(rec)

    if customer:
        customer.outstanding_balance = float(customer.outstanding_balance or 0) - refund_amount

    db.commit()
    db.refresh(rec)
    return _return_to_out(rec, db)


# ---------------------------------------------------------------------------
# 세금계산서 라우트
# ---------------------------------------------------------------------------
def _invoice_to_out(inv: TaxInvoice, db: Session) -> TaxInvoiceOut:
    order = db.query(SalesOrder).filter(SalesOrder.id == inv.order_id).first()
    customer_name = None
    if order:
        customer = db.query(Customer).filter(Customer.id == order.customer_id).first()
        customer_name = customer.name if customer else None
    return TaxInvoiceOut(
        id=inv.id, order_id=inv.order_id, customer_name=customer_name,
        invoice_no=inv.invoice_no, issue_date=inv.issue_date,
        supply_amount=float(inv.supply_amount), vat_amount=float(inv.vat_amount),
        total_amount=float(inv.total_amount),
    )


@app.get("/tax-invoices", response_model=list[TaxInvoiceOut], tags=["세금계산서"])
def list_invoices(db: Session = Depends(get_db)):
    items = db.query(TaxInvoice).order_by(TaxInvoice.id.desc()).all()
    return [_invoice_to_out(i, db) for i in items]


@app.post("/tax-invoices", response_model=TaxInvoiceOut, tags=["세금계산서"])
def create_invoice(payload: TaxInvoiceIn, db: Session = Depends(get_db)):
    order = db.query(SalesOrder).filter(SalesOrder.id == payload.order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="주문을 찾을 수 없습니다")
    existing = db.query(TaxInvoice).filter(TaxInvoice.order_id == payload.order_id).first()
    if existing:
        raise HTTPException(status_code=400, detail="이미 세금계산서가 발행된 주문입니다")

    inv = TaxInvoice(
        order_id=order.id, invoice_no=payload.invoice_no, issue_date=payload.issue_date,
        supply_amount=order.supply_amount, vat_amount=order.vat_amount, total_amount=order.total_amount,
    )
    db.add(inv)
    db.commit()
    db.refresh(inv)
    return _invoice_to_out(inv, db)


# ---------------------------------------------------------------------------
# 출고 리포트 (일별 / 월별 출고 금액 및 마진율)
# 주의: 출고완료 상태인 주문의 shipped_at(출고 처리 시각) 기준으로 집계한다.
# 반품은 별도 집계이며, 이 리포트에서는 아직 차감하지 않는다.
# ---------------------------------------------------------------------------
def _summarize_orders(orders: list[SalesOrder], period_label: str) -> ShipmentReportOut:
    supply = sum(float(o.supply_amount) for o in orders)
    cost = sum(float(o.cost_amount) for o in orders)
    margin = supply - cost
    margin_rate = round(margin / supply * 100, 2) if supply else None
    return ShipmentReportOut(
        period=period_label, order_count=len(orders),
        supply_amount=supply, cost_amount=cost,
        margin_amount=margin, margin_rate=margin_rate,
    )


@app.get("/reports/shipments/daily", response_model=ShipmentReportOut, tags=["출고리포트"])
def daily_shipment_report(target_date: date, db: Session = Depends(get_db)):
    orders = db.query(SalesOrder).filter(
        SalesOrder.status == "출고완료",
        SalesOrder.shipped_at >= datetime.combine(target_date, datetime.min.time()),
        SalesOrder.shipped_at < datetime.combine(target_date, datetime.max.time()),
    ).all()
    return _summarize_orders(orders, target_date.isoformat())


@app.get("/reports/shipments/monthly", response_model=ShipmentReportOut, tags=["출고리포트"])
def monthly_shipment_report(year: int, month: int, db: Session = Depends(get_db)):
    last_day = monthrange(year, month)[1]
    start = datetime(year, month, 1)
    end = datetime(year, month, last_day, 23, 59, 59)
    orders = db.query(SalesOrder).filter(
        SalesOrder.status == "출고완료",
        SalesOrder.shipped_at >= start,
        SalesOrder.shipped_at <= end,
    ).all()
    return _summarize_orders(orders, f"{year}-{month:02d}")


@app.get("/api/health")
def health_check():
    return {"status": "ok", "message": "Sejun ERP API가 정상적으로 실행 중입니다."}


# 브라우저에서 서버 주소(예: http://192.168.0.5:8000)로 바로 접속하면
# frontend/index.html을 따로 열지 않아도 화면이 뜨도록, 여기서 직접 서빙한다.
# (반드시 다른 라우트들 뒤에 마운트해야 /products 같은 API 경로를 가리지 않는다.)
app.mount("/", StaticFiles(directory="static", html=True), name="static")
